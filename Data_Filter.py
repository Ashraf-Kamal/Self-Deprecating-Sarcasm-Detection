# -*- coding: utf-8 -*-
import warnings
warnings.filterwarnings("ignore")

import tweepy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.compat import range
import pandas as pd
import os
import csv
import spacy
from spacy import displacy
import re
from spacy.lang.en import English

import en_core_web_sm
nlp = en_core_web_sm.load()
nlp = spacy.load('en')
parser = English()

def print_fine_pos(token):
    return (token.pos_)

def print_fine_tag(token):
    return (token.tag_)
	
'''
This program is implemented in Python 2.7. In this program, spaCy, a popular open-source library for advanced NLP tasks in Python is used.
It recognizes linguistic markers like punctuation, and it has the ability to split these punctuation tokens from word tokens.
Hence, the performance of the Parts-Of-Speech (POS) tagger using spaCy is not affected, if it is applied to cleaned data. Using spaCy, tweets tokenization 
and POS tagging task are accomplished. Thereafter, explicit self-referential tweets-based 'Specific' and 'Generic' patterns are matched. 
These patterns are mainly based on either tokens or the sequential order of tags and tokens or vice-versa. 
'''

file_exists = os.path.isfile('Filtered_Data_Sets\Twitter-280\Explicit\Sarcasm_Twitter-280.csv')
file_exists1 = os.path.isfile('Filtered_Data_Sets\Twitter-280\Implicit\Sarcasm_Twitter-280.csv')

csvFile = open('Filtered_Data_Sets\Twitter-280\Explicit\Sarcasm_Twitter-280.csv', 'ab')
fields = ('Tweet_Text','Tweet_Id') #field names
csvWriter = csv.writer(csvFile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
csvWriter = csv.DictWriter(csvFile, fieldnames=fields)
if file_exists:
   csvWriter.writeheader()

csvFile1 = open('Filtered_Data_Sets\Twitter-280\Implicit\Sarcasm_Twitter-280.csv', 'ab')
fields1 = ('Tweet_Text','Tweet_Id') #field names
csvWriter1 = csv.writer(csvFile1, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
csvWriter1 = csv.DictWriter(csvFile1, fieldnames=fields1)
if file_exists1:
    csvWriter1.writeheader()   

def Get_Tag(text):
  sents = nlp(text.decode('utf-8'))
  tags=[]
  for tok in sents:  
    tags.append((tok.text,print_fine_pos(tok),print_fine_tag(tok)))    
  		       
  return tags

def Filter_Preprocess(tweet_id,text,tweet):
  for i in range(len(tweet)):    
    if re.search(r"(\bUH\b)", tweet[i][2]) and re.search(r"(\bi\b|\bmy\b|\bwe\b|\bare\b)", tweet[i+1][0]):
    	csvWriter.writerow({'Tweet_Text': text.encode('utf-8').replace('\n', '').replace('\r', ' ').replace('\\', ' ')})
    	print '1'
    	return 0

    elif re.search(r"(\bwe\b|\bi\b)", tweet[i][0]) and re.search(r"(\blove\b)", tweet[i+1][0]) and re.search(r"(\bit\b|\bwhen\b)", tweet[i+2][0]):
    	csvWriter.writerow({'Tweet_Text': text.encode('utf-8').replace('\n', '').replace('\r', ' ').replace('\\', ' ')})
    	print '2'
    	return 0    

    elif re.search(r"(\bwhen\b)", tweet[i][0]) and re.search(r"(\bmy\b|\bour\b)", tweet[i+1][0]):
    	csvWriter.writerow({'Tweet_Text': text.encode('utf-8').replace('\n', '').replace('\r', ' ').replace('\\', ' ')})
    	print '3'
    	return 0

    elif re.search(r"(\bam\b|\bare\b)", tweet[i][0]) and re.search(r"(\bstill\b)", tweet[i+1][0]):
    	csvWriter.writerow({'Tweet_Text': text.encode('utf-8').replace('\n', '').replace('\r', ' ').replace('\\', ' ')})
    	print '4'
    	return 0

    elif re.search(r"(\bmyself\b|\bourself\b)", tweet[i][0]) and re.search(r"(\bJJ\b|\bRB\b)", tweet[i+1][0]):
    	csvWriter.writerow({'Tweet_Text': text.encode('utf-8').replace('\n', '').replace('\r', ' ').replace('\\', ' ')})
    	print '5'
    	return 0

    elif re.search(r"(\boh\b|\bwow\b|\byeah\b)", tweet[i][0]) and re.search(r"(\bi\b|\bwe\b)", tweet[i+1][0]) and re.search(r"(\breal+y\b|\bgreat+\b)", tweet[i+2][0]):
    	csvWriter.writerow({'Tweet_Text': text.encode('utf-8').replace('\n', '').replace('\r', ' ').replace('\\', ' ')})
    	print '6'
    	return 0

    elif re.search(r"(\bi\b|\bwe\b)", tweet[i][0]) and re.search(r"(\bMD\b|\bRB\b)", tweet[i+1][0]):
    	csvWriter.writerow({'Tweet_Text': text.encode('utf-8').replace('\n', '').replace('\r', ' ').replace('\\', ' ')})
    	print '7'
    	return 0

    elif re.search(r"(\bi\b|\bmy\b|\bme\b|\bmine\b|\bmyself\b|\bam\b|\bwe\b|\bus\b|\bour\b|\bourselves\b|\bare\b)", tweet[i][0]):
    	csvWriter.writerow({'Tweet_Text': text.encode('utf-8').replace('\n', '').replace('\r', ' ').replace('\\', ' ')})
    	print '8'
    	return 0

  csvWriter1.writerow({'Tweet_Text': text.replace('\n', '').replace('\r', ' ').replace('\\', ' ')})
  print 'bye'

count=0 # To count no. of tweets
try:
   df_new = pd.read_csv('Pre_Processed_Data_Sets\Preprocess_2\Twitter-280\Sarcasm_Twitter-280.csv')
   writer = pd.ExcelWriter('Pre_Processed_Data_Sets\Preprocess_2\Twitter-280\Sarcasm_Twitter-280.xlsx')
   df_new.to_excel(writer, index = False)
   writer.save()
   wb = load_workbook('Pre_Processed_Data_Sets\Preprocess_2\Twitter-280\Sarcasm_Twitter-280.xlsx')
   #wb = load_workbook('Data-Set\Short.xlsx')
   sheet = wb["Sheet1"]
   rows = sheet.max_row
   column = sheet.max_column
   rows=rows+1
   tweets=[]   
    # Append tweets fetch after pre-processing
   for i in range(2, rows):  
      text=(sheet.cell(row=i, column=1).value)
      #print text
      tweet_id=(sheet.cell(row=i, column=2).value)
      if text!=None and len(text.split())>3: 
        Filter_Preprocess(tweet_id,text.encode('utf-8').replace('\\', ' '),Get_Tag(text.encode('utf-8').replace('\\', ' ')))
        count +=1
        #print count
        rows=rows+1
      else:
        count +=1
        rows=rows+1
   print 'Done', count
    
except tweepy.TweepError:
    print("Error while loading raw data")   
   


    


          

