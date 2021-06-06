# -*- coding: utf-8 -*-
# encoding=utf8
import sys
reload(sys)
sys.setdefaultencoding('utf8')
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.compat import range
import pandas as pd
import os
import Data_Preprocess as DPF

#This program is implemented in Python 2.7. It takes tweet as an input to perform data cleaning tasks using DFS.Preprocessing_Start function.

count=0 # To count no. of tweets
wb = load_workbook('Raw_Data_Sets\Twitter-280\Sarcasm_Twitter-280.xlsx')
sheet = wb["Sheet1"]
rows = sheet.max_row
column = sheet.max_column
rows=rows+1
tweets=[] 
tweets_id=[]
#tweets_label=[]   
#Append tweets fetch after pre-processing.
try:	
	for i in range(2, rows):
	  Tweet_Id=(sheet.cell(row=i, column=1).value)
	  text=(sheet.cell(row=i, column=2).value)      
		  #label=(sheet.cell(row=i, column=1).value)
	  print text 
	  if text!=None and len(text.split())>3:
		tweets.append(DPF.Preprocessing_Start(text.encode('utf-8').replace('\\', ' ').decode('unicode_escape').encode('ascii','ignore').strip()))
		tweets_id.append(str(Tweet_Id))
		#tweets_label.append(label)
		count +=1
		rows=rows+1

	df = {'Tweet_Id': tweets_id, 'Tweet_Text': tweets}
	tweets=pd.DataFrame(df)
	tweets.to_csv('Pre_Processed_Data_Sets\Preprocess1_CSV\Twitter-280\Sarcasm_Twitter-280.csv', sep=',', index=False, encoding='utf-8')

	#Below code is to convert CSV file to Excel file.Sem-Eval
	df_new = pd.read_csv('Pre_Processed_Data_Sets\Preprocess_CSV\Twitter-280\Sarcasm_Twitter-280.csv')
	writer = pd.ExcelWriter('Pre_Processed_Data_Sets\Preprocess1_Excel\Twitter-280\Sarcasm_Twitter-280.xlsx')
	df_new.to_excel(writer, index = False)
except tweepy.TweepError:
    print("Failed to fetched tweet against the tweet id...")
writer.save()

