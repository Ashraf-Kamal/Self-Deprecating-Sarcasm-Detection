#!/usr/bin/python3
# -*- coding: utf-8 -*-
#coding=utf-8
import tweepy
from openpyxl import load_workbook
import os
import sys
import string
import simplejson
import xlrd
import csv #Import csv
from xlrd import open_workbook
import pandas as pd

#This program is implemented in Python 2.7. In this program, it takes tweet ids as an input for a dataset, 
#and tweets are fetched against tweet ids respectively using Twitter REST API.

auth = tweepy.OAuthHandler('xxxxxxxxxxxxxxxxxx','yyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyyy')
auth.set_access_token('ZZZZZZZZZZZZ-zzzzzzzzzzzzz','llllllllllllllllllllllllllllllllllllll')
api = tweepy.API(auth)
response_ids = []

wb = load_workbook(filename='Twitter_Id\Twitter-280\TweetId.xlsx')
ws = wb['Sheet1']
ws = wb.active
col = ws['A']

for row in ws.iter_rows():
    for col in row:
        response_ids.append(col.value)
file_exists = os.path.isfile('Raw_Data_Sets\Twitter-280\Sarcasm_Twitter-280.csv')
# How many elements each 
# list should have 
n = 100 
# using list comprehension 
final = [response_ids[i * n:(i + 1) * n] for i in range((len(response_ids) + n - 1) // n )] 
# Open/create a file to append data to 407234283049201664
csvFile = open('Raw_Data_Sets\Twitter-280\Sarcasm_Twitter-280.csv', 'ab')
fields = ('Tweet_Id', 'Tweet_Text') #field names
csvWriter = csv.DictWriter(csvFile, fieldnames=fields)

try:
	if not file_exists:
		csvWriter.writeheader()
	csvWriter = csv.writer(csvFile)
	if not file_exists:
		csvWriter.writeheader()	
	for x in final:
		tweets = api.statuses_lookup(id_=x)
		for tweet in tweets:              
			csvWriter.writerow({'Tweet_Id': tweet.id_str, 'Tweet_Text': tweet.text.encode('utf-8').replace('\n', '').replace('\r', ' ')})       
except tweepy.TweepError:
    print("Failed to fetched tweet against the tweet id...")

csvFile.close()


